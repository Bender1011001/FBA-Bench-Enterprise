from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define styles
currency_style = NamedStyle(name="currency", number_format="$#,##0.00")
percent_style = NamedStyle(name="percent", number_format="0.00%")
decimal_style = NamedStyle(name="decimal", number_format="0.00")

wb.add_named_style(currency_style)
wb.add_named_style(percent_style)
wb.add_named_style(decimal_style)

# Inputs sheet
ws_inputs = wb.create_sheet("Inputs")

# Headers
headers = ["Field", "Value", "Unit", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws_inputs.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Data
input_data = [
    ["Current_eval_hours_per_month", 500, "number", "Current monthly evaluation hours (placeholder)"],
    ["Hourly_cost", 100, "number", "USD per hour (fully loaded labor cost)"],
    ["Current_tool_cost_per_month", 20000, "number", "Current monthly tool subscription cost"],
    ["Error_rate", 0.15, "percent", "Current error rate in evaluations (%)"],
    ["Revenue_at_risk_per_error", 50000, "number", "Revenue at risk per error (USD)"],
    ["Expected_hours_reduction_pct", 0.6, "percent", "Expected % reduction in hours (e.g., 60%)"],
    ["Expected_tool_cost_reduction_pct", 0.5, "percent", "Expected % reduction in tool costs (e.g., 50%)"],
    ["Expected_error_reduction_pct", 0.8, "percent", "Expected % reduction in errors (e.g., 80%)"],
    ["Revenue_uplift_per_month", 100000, "number", "Expected monthly revenue uplift (USD)"],
    ["Investment_cost", 50000, "number", "One-time or annual investment cost (USD)"]
]

for row_idx, row_data in enumerate(input_data, start=2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_inputs.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 2:  # Value column
            if "pct" in row_data[0]:
                cell.style = percent_style
            elif "cost" in row_data[0] or "Revenue" in row_data[0]:
                cell.style = currency_style
        elif col_idx == 3:  # Unit column
            if "percent" in str(value):
                cell.value = "percent"

# Results sheet
ws_results = wb.create_sheet("Results")

# Headers
result_headers = ["Metric", "Value", "Formula/Notes"]
for col, header in enumerate(result_headers, 1):
    cell = ws_results.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Data and formulas (row numbers based on Inputs: B2 to B11)
result_data = [
    ["Hours_saved", "=Inputs!B2*Inputs!B6", "Current_eval_hours_per_month * Expected_hours_reduction_pct"],
    ["Labor_savings", "=B2*Inputs!B3", "Hours_saved * Hourly_cost (monthly)"],
    ["Tool_savings", "=Inputs!B4*Inputs!B7", "Current_tool_cost_per_month * Expected_tool_cost_reduction_pct (monthly)"],
    ["Error_cost_avoidance", "=(Inputs!B2*Inputs!B4)*Inputs!B5*Inputs!B8", "(Current_eval_hours_per_month * Error_rate) * Revenue_at_risk_per_error * Expected_error_reduction_pct (monthly)"],
    ["Revenue_uplift_per_month", "=Inputs!B9", "From Inputs (monthly)"],
    ["Total_savings", "=SUM(B2:B6)", "Sum of monthly savings components"],
    ["Investment_cost", "=Inputs!B10", "From Inputs"],
    ["Net_benefit", "=B7-B9", "Total_savings - Investment_cost (monthly net)"],
    ["ROI_pct", "=IF(B9=0, \"\", B8/B9)", "Net_benefit / Investment_cost"],
    ["Payback_months", "=IF(B8<=0, \"\", B9 / (B7/12))", "Investment_cost / (Total_savings/12)"]
]

for row_idx, row_data in enumerate(result_data, start=2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_results.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 2 and isinstance(value, str) and "=" in value:  # Formula cells
            # Apply styles based on metric
            if any(term in row_data[0] for term in ["savings", "cost", "benefit"]):
                cell.style = currency_style
            elif "ROI_pct" in row_data[0]:
                cell.style = decimal_style
            elif "Payback" in row_data[0]:
                cell.style = decimal_style
        elif col_idx == 3:
            cell.alignment = Alignment(wrap_text=True)

# Sample sheet with prefilled examples
ws_sample = wb.create_sheet("Sample")
ws_sample.append(["Sample Inputs - Use these as starting point for demos (placeholders only)"])
ws_sample.append([])  # Empty row

# Copy inputs but with sample values
sample_inputs = [
    ["Current_eval_hours_per_month", 500, "hours/month"],
    ["Hourly_cost", 100, "$/hour"],
    ["Current_tool_cost_per_month", 20000, "$/month"],
    ["Error_rate", 15, "%"],
    ["Revenue_at_risk_per_error", 50000, "$"],
    ["Expected_hours_reduction_pct", 60, "%"],
    ["Expected_tool_cost_reduction_pct", 50, "%"],
    ["Expected_error_reduction_pct", 80, "%"],
    ["Revenue_uplift_per_month", 100000, "$/month"],
    ["Investment_cost", 50000, "$"]
]

for row_data in sample_inputs:
    ws_sample.append(row_data)

# Note
ws_sample.append([])
ws_sample.append(["Note: These are example values. Results sheet will compute based on these inputs. No real data used."])

# Adjust column widths
for ws in [ws_inputs, ws_results, ws_sample]:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_path = "repos/fba-bench-enterprise/marketing/tools/roi_calculator.xlsx"
wb.save(output_path)
print(f"ROI Calculator generated at {output_path}")
